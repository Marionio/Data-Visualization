# Import the necessary libraries
import openpyxl as xl
import pandas as pd
from math import log, floor

data = pd.read_csv('CMaster.csv')
df = pd.DataFrame(data)
pd.options.display.max_columns = 20

# Fixing main df 
df['TOTAL_TIME'] = pd.to_timedelta(df['TOTAL_TIME'])
df['TIME_TO_PICKUP'] = pd.to_timedelta(df['TIME_TO_PICKUP'])
df['TIME_TO_DELIVER'] = pd.to_timedelta(df['TIME_TO_DELIVER'])

# Long number formatter
def human_format(number):
    units = ['', 'K', 'M', 'G', 'T', 'P']
    k = 1000.0
    magnitude = int(floor(log(number, k)))
    return '%.2f%s' % (number / k**magnitude, units[magnitude])

# Labels info
# Total Deliveries
tot_deliveries =df['STORE_NAME'].count()
# Total Items
tot_items = df['TOTAL_ITEM_COUNT'].sum()
# Total subtotals
tot_dls = df['SUBTOTAL_IN_DOLLARS'].sum()
# Percentage of orders delivered
perf = df['ORDER_STATUS'].value_counts(normalize=True) * 100
perf = perf.iloc[0]
# Top 5 fastest deliveries
fastest_delivery = df.loc[:,('TIME_TO_PICKUP','TIME_TO_DELIVER','TOTAL_TIME','STORE_NAME')]
fastest_delivery= fastest_delivery.rename(columns = {'TIME_TO_PICKUP':'PICKUP','TIME_TO_DELIVER':'DELIVER','TOTAL_TIME':'TOTAL', 'STORE_NAME':'VENDOR'})
fastest_delivery = fastest_delivery.sort_values(by='TOTAL')
top5 = fastest_delivery.head(10)
top5 = pd.DataFrame(top5)
# Top highest subtotals
most_expensive = df[['STORE_NAME','SUBTOTAL_IN_DOLLARS']]
most_expensive= most_expensive.rename(columns={'SUBTOTAL_IN_DOLLARS':'SUBTOTAL', 'STORE_NAME':'VENDOR'})
most_expensive= most_expensive.sort_values(by='SUBTOTAL', ascending=False)
most_expensive = most_expensive.head(10)
# Longest time to deliver
slowest_deliveries = df.loc[:,('TIME_TO_DELIVER','STORE_NAME')]
slowest_deliveries = slowest_deliveries.rename(columns = {'TIME_TO_DELIVER':'DELIVER','STORE_NAME':'VENDOR'})
slowest_deliveries = slowest_deliveries.sort_values(by='DELIVER', ascending=False)
slow5 = slowest_deliveries.head(5)
slow5 = pd.DataFrame(slow5)
# Top 10 Vendors
top_10 = data['STORE_NAME'].value_counts().head(10).sort_values()
top_10 = pd.DataFrame(top_10)

# Price range
prc= df[['STORE_NAME','SUBTOTAL_IN_DOLLARS']]
prc = prc[(prc['SUBTOTAL_IN_DOLLARS']<100)]
prc = prc['SUBTOTAL_IN_DOLLARS'].squeeze()
prc= prc.value_counts(bins=10).sort_index()
prc = pd.DataFrame(prc)
# Time DataFrame
hourdf = df['ORDER_CREATED_TIME']
hourdf = pd.to_datetime(hourdf, format = '%Y/%m/%d %H:%M:%S')
hourdf = hourdf.dt.hour
hourdf = pd.DataFrame(hourdf)
hourdf = hourdf.value_counts(subset = 'ORDER_CREATED_TIME').sort_index()
hourdf= pd.DataFrame(hourdf)

thetop = top_10.index.tolist()
thetopdf= df[df['STORE_NAME'].isin(thetop)]
thetopdf = thetopdf.groupby(by='STORE_NAME')
thetopdf = thetopdf['TOTAL_TIME'].mean()
thetopdf = thetopdf.sort_values()
thetopdf = pd.DataFrame(thetopdf)

# Report data
top_vendor = thetop[-1]
top_vendor_deliveries = top_10.tail(1).values
time_top_vendor = df[(df['STORE_NAME'] == top_vendor)]
time_top_vendor = time_top_vendor['TIME_TO_DELIVER'].sum()
first_delivery = df['ACTUAL_DELIVERY_TIME'].min()
first_delivery = first_delivery[0:11]
tot_time_driving = df['TIME_TO_DELIVER'].sum()
min_subtotal= df['SUBTOTAL_IN_DOLLARS'].min()
max_subtotals= df['SUBTOTAL_IN_DOLLARS']
max_subtotals= max_subtotals[(max_subtotals <100)]
max_subtotals= max_subtotals.max()
avg_delivery_fix = df[(df['TOTAL_TIME'] < '1:20:00')]
avg_delivery_fix = avg_delivery_fix['TOTAL_TIME'].mean()
avg_delivery_time = avg_delivery_fix.components

Dasher = input('Dasher name:')

# Fixing data type

top5['TOTAL'] = pd.to_timedelta(top5['TOTAL'])
top5['PICKUP'] = pd.to_timedelta(top5['PICKUP'])
top5['DELIVER'] = pd.to_timedelta(top5['DELIVER'])

slow5['DELIVER'] = pd.to_timedelta(top5['DELIVER'])


# Writing tables to sheet
dfs = [top_10,thetopdf,prc,hourdf]
startcol = 0
with pd.ExcelWriter('Report.xlsx') as writer:
    top5.to_excel(writer, sheet_name='Report' ,engine="openpyxl",index =False, startcol=27, startrow= 30)
    most_expensive.to_excel(writer, sheet_name='Report' ,engine="openpyxl",index =False, startcol=33, startrow= 30)
    for df in dfs:
            df.to_excel(writer, sheet_name='Data',engine="openpyxl", startcol=startcol)
            try:
                startcol += (df.shape[1] + 2)
            except(IndexError):
                pass
        
    
# Start workbook
wb = xl.load_workbook('Report.xlsx')
R_S = wb.active
D_S = wb['Data']
for r in range (30,42):
    R_S[f'AB{r}'].number_format = 'mm:ss'
    R_S[f'AC{r}'].number_format = 'mm:ss'
    R_S[f'AD{r}'].number_format = 'mm:ss'
    R_S[f'AI{r}'].number_format = '#,##,0.00$'
R_S.column_dimensions['AH'].bestFit
for r in range (2,12):
    D_S[f'E{r}'].number_format = 'mm:ss'
    D_S[f'G{r}'] = D_S[f'G{r}'].value.replace("]","$")
    D_S[f'G{r}'] = D_S[f'G{r}'].value.replace("(","")
    D_S[f'G{r}'] = D_S[f'G{r}'].value.replace(",","$-")
    D_S[f'G{r}'].number_format = '#,##,0.00$'
    
# Labels
R_S.merge_cells('A10:I15')
cell = R_S.cell(row=10, column=1)  
cell.value = 'DoorDash Summary'
cell.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
cell.font = xl.styles.Font(size= 20)

R_S.merge_cells('A16:I20')
cell = R_S.cell(row=16, column=1)  
cell.value = 'By: Mario Escobar'
cell.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
cell.font = xl.styles.Font(size= 20)
R_S.merge_cells('K3:P6')  
  
cell2 = R_S.cell(row=3, column=11)  
cell2.value = 'Outline'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 20)

R_S.merge_cells('K9:P10')  
  
cell2 = R_S.cell(row=9, column=11)  
cell2.value = 'Introduction'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('K13:P14')  
  
cell2 = R_S.cell(row=13, column=11)  
cell2.value = 'Overall Summary'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('K17:P18')  
  
cell2 = R_S.cell(row=17, column=11)  
cell2.value = 'Additional Charts'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('K21:P22')  
  
cell2 = R_S.cell(row=21, column=11)  
cell2.value = 'Conclusion'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('T3:Y5')  

cell2 = R_S.cell(row=3, column=20)  
cell2.value = 'Introduction'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('T8:Z40')

cell2 = R_S.cell(row=8, column=20)
cell2.value = f"Greeting, {Dasher}.\nEver since you started your first delivery on {first_delivery} you have been collecting data for DoorDash. \nTime goes fast for {human_format(tot_deliveries)} deliveries. Did you know you have spent at least {time_top_vendor}h driving to {top_vendor} customers?\n\nNext are the total subtotals you have handled for customers, how many items you’ve handled, your top 10 vendors, fastest deliveries and highest subtotals for deliveries.\n\nThe charts in the following page are the average times it takes from a customer to receive an order from the top 10 vendors (1), hours of the day in which orders were placed by customers(2) and finally a chart counting orders from subtotals as little as {min_subtotal}$ to {max_subtotals}$(3)."
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='top',wrapText=True)
cell2.font = xl.styles.Font(size= 16)

R_S.merge_cells('AB1:AE7')  

cell = R_S.cell(row=1, column=28)  
cell.value = f'{human_format(tot_deliveries)} Deliveries'  
cell.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
cell.font = xl.styles.Font(size= 20)

R_S.merge_cells('AF1:AJ7')  
  
cell2 = R_S.cell(row=1, column=32)  
cell2.value = f'${human_format(tot_dls)} Subtotals Handled'  
cell2.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('AB8:AE14')  
  
cell3 = R_S.cell(row=8, column=28)  
cell3.value = f'{human_format(tot_items)} Items Handled'  
cell3.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
cell3.font = xl.styles.Font(size= 20)

R_S.merge_cells('AF8:AJ14')  
  
cell4 = R_S.cell(row=8, column=32)  
cell4.value = f'Orders Status {int(perf)}% Delivered'
cell4.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
cell4.font = xl.styles.Font(size= 18)

R_S.merge_cells('AB30:AE30')  
  
cell5 = R_S.cell(row=30, column=28)  
cell5.value = 'Fastest deliveries'  
cell5.alignment = xl.styles.Alignment(vertical='center', horizontal = 'center')
cell5.font = xl.styles.Font(size= 16)

R_S.merge_cells('AG30:AJ30')  
  
cell6 = R_S.cell(row=30, column=33)  
cell6.value = 'Highest Subtotals'  
cell6.alignment = xl.styles.Alignment(vertical='center', horizontal = 'center')
cell6.font = xl.styles.Font(size= 16)


R_S.merge_cells('AU3:BA5')  

cell2 = R_S.cell(row=3, column=47)  
cell2.value = 'Conclusion'  
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='center')
cell2.font = xl.styles.Font(size= 18)

R_S.merge_cells('AU8:BA40')

cell2 = R_S.cell(row=8, column=47)  
cell2.value = f"Your average delivery time is {str(avg_delivery_time.minutes)+':'+str(avg_delivery_time.seconds)} from the time the order was placed.\n\nYou have been on the road for for at least {tot_time_driving}h.\n\nWe spend much more time on the road than the average person. Remember to always drive safely.\n\nYou have handled {human_format(tot_dls)} dlls but the most valuable thing is you.\n\nGood luck moving forward, {Dasher}.\n\n\nSincerely, Mario"
cell2.alignment = xl.styles.Alignment(horizontal='left', vertical='top',wrapText=True)
cell2.font = xl.styles.Font(size= 16)


# Chartmaking
barchart = xl.chart.BarChart()
data = xl.chart.Reference(D_S, min_col=2,max_col=2,min_row =1 ,max_row=12)
categories = xl.chart.Reference(D_S, min_col=1,max_col=1,min_row =2 ,max_row=12)
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
barchart.title = "Top 10 Vendors"
barchart.type = "bar"
barchart.legend = None
s= barchart.series[0]
s.graphicalProperties.solidFill = "ff0000" 
R_S.add_chart(barchart,"AB15")

# Chart 2, times avg
barchart2 = xl.chart.BarChart()
data2 = xl.chart.Reference(D_S, min_col=5,max_col=5,min_row =1 ,max_row=12)
categories2 = xl.chart.Reference(D_S, min_col=4,max_col=4,min_row =2 ,max_row=12)
barchart2.add_data(data2, titles_from_data=True)
barchart2.set_categories(categories2)
barchart2.title = "Top 10 Average times to complete an order"
barchart2.legend = None
barchart2.y_axis.majorUnit = 0.003472
s2= barchart2.series[0]
s2.graphicalProperties.solidFill = "ff0000" 
R_S.add_chart(barchart2,"AK1")

# Chart 3, times of day
barchart3 = xl.chart.BarChart()
data3 = xl.chart.Reference(D_S, min_col=11,max_col=11,min_row =1 ,max_row=25)
categories3 = xl.chart.Reference(D_S, min_col=10,max_col=10,min_row =2 ,max_row=25)
barchart3.add_data(data3, titles_from_data=True)
barchart3.set_categories(categories3)
barchart3.title = "Hours ordered at"
barchart3.legend = None
s3= barchart3.series[0]
s3.graphicalProperties.solidFill = "ff0000" 
R_S.add_chart(barchart3,"AK16")

# Chart 4 Price Range

barchart4 = xl.chart.BarChart()
data4 = xl.chart.Reference(D_S, min_col=8,max_col=8,min_row =1 ,max_row=12)
categories4 = xl.chart.Reference(D_S, min_col=7,max_col=7,min_row =2 ,max_row=12)
barchart4.add_data(data4, titles_from_data=True)
barchart4.set_categories(categories4)
barchart4.title = "Order Subtotals (excludes tip and tax)"
barchart4.legend = None
s4= barchart4.series[0]
s4.graphicalProperties.solidFill = "ff0000" 
R_S.add_chart(barchart4,"AK31")

# Save Workbook
wb.save(f'{Dasher}_Report.xlsx')

